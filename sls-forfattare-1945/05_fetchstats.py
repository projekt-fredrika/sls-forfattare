#!/usr/bin/env python3
"""
Script to fetch Wikipedia pageview statistics for authors from the SPARQL query.
Fetches pageview data for the last 12 months for Swedish, Finnish, and English Wikipedia.
Outputs results to an Excel file.
"""

import requests
import time
import json
from datetime import datetime, timedelta
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def run_sparql_query(query):
    """Execute SPARQL query and return results"""
    url = "https://query.wikidata.org/sparql"
    
    headers = {
        'User-Agent': 'SLS-Forfattare-1917/1.0 (projektfredrika.fi)',
        'Accept': 'application/sparql-results+json'
    }
    
    params = {
        'query': query,
        'format': 'json'
    }
    
    print("Running SPARQL query...")
    response = requests.get(url, params=params, headers=headers, timeout=30)
    response.raise_for_status()
    
    data = response.json()
    bindings = data.get('results', {}).get('bindings', [])
    
    print(f"Found {len(bindings)} results")
    return bindings


def get_wikipedia_pageviews(page_title, language, start_date, end_date):
    """
    Fetch Wikipedia pageview statistics for a given page title.
    
    Args:
        page_title (str): Wikipedia page title
        language (str): Language code (sv, fi, en)
        start_date (str): Start date in YYYYMMDD format
        end_date (str): End date in YYYYMMDD format
    
    Returns:
        int: Total pageviews for the period
    """
    if not page_title:
        return 0
    
    # Wikipedia Pageviews API
    # https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article/{project}/{access}/{agent}/{article}/{granularity}/{start}/{end}
    
    project = f"{language}.wikipedia"
    access = "all-access"
    agent = "all-agents"
    article = page_title.replace(" ", "_")
    granularity = "monthly"  # or daily
    
    url = f"https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article/{project}/{access}/{agent}/{article}/{granularity}/{start_date}/{end_date}"
    
    try:
        headers = {
            'User-Agent': 'SLS-Forfattare-1917/1.0 (projektfredrika.fi)',
            'Accept': 'application/json'
        }
        
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 404:
            # Page doesn't exist
            return 0
        
        response.raise_for_status()
        data = response.json()
        
        # Sum up all pageviews
        total_views = 0
        items = data.get('items', [])
        for item in items:
            total_views += item.get('views', 0)
        
        return total_views
    
    except Exception as e:
        print(f"  Error fetching pageviews for '{page_title}' in {language}: {str(e)}")
        return 0


def get_wikipedia_article_length(page_title, language):
    """
    Fetch Wikipedia article length (number of bytes) for a given page title.
    Uses the Action API with prop=info which returns a reliable `length` field
    and handles redirects.

    Args:
        page_title (str): Wikipedia page title
        language (str): Language code (sv, fi, en)

    Returns:
        int: Page length in bytes, 0 if not available
    """
    if not page_title:
        return 0

    url = f"https://{language}.wikipedia.org/w/api.php"
    params = {
        'action': 'query',
        'format': 'json',
        'formatversion': '2',
        'prop': 'info',
        'titles': page_title,
        'redirects': '1',
    }

    try:
        headers = {
            'User-Agent': 'SLS-Forfattare-1917/1.0 (projektfredrika.fi)',
            'Accept': 'application/json'
        }
        response = requests.get(url, headers=headers, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        pages = data.get('query', {}).get('pages', [])
        if not pages:
            return 0
        page = pages[0]
        if page.get('missing'):
            return 0
        return int(page.get('length', 0) or 0)
    except Exception as e:
        print(f"  Error fetching length for '{page_title}' in {language}: {str(e)}")
        return 0


def get_date_range_one_year_back():
    """Get start and end dates for the past 12 months"""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    
    # Format as YYYYMMDD
    return start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d")


def clean_title(title):
    """Remove Wikipedia namespace prefixes if present"""
    if title:
        # Remove common prefixes
        for prefix in ['User:', 'Talk:', 'File:', 'Category:', 'Template:']:
            if title.startswith(prefix):
                return title[len(prefix):].strip()
    return title


def extract_qcode(item_url):
    """Extract Q-code from Wikidata entity URL"""
    if not item_url:
        return ''
    # Extract Q-code from URL like http://www.wikidata.org/entity/Q123456
    # or https://www.wikidata.org/entity/Q123456
    if '/entity/' in item_url:
        qcode = item_url.split('/entity/')[-1]
        # Remove any trailing characters or fragments
        qcode = qcode.split('#')[0]
        return qcode
    return item_url  # Return as-is if not a standard URL


def create_excel(rows, output_file):
    """Create Excel file with SPARQL results and pageview statistics"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Author Statistics"
    
    if not rows:
        wb.save(output_file)
        return
    
    # Get headers from all rows to ensure all columns are included
    # Collect all unique keys from all rows
    all_keys = set()
    for row_data in rows:
        all_keys.update(row_data.keys())
    
    # Define column order as they appear in the SPARQL query SELECT statement
    sparql_column_order = [
        'item',
        'dob',
        'itemLabel_sv',
        'itemLabel_fi',
        'itemLabel_en',
        'itemDescription_sv',
        'itemDescription_fi',
        'itemDescription_en',
        'forfattare_sida',
        'forfattare_ref',
        'wp_sv_title',
        'wp_fi_title',
        'wp_en_title'
    ]
    
    # Excluded keys (commented-out fields)
    excluded_keys = ['pobLabel', 'dod', 'podLabel']
    
    headers = []
    
    # Add columns in SPARQL query order, if they exist in the data
    for col in sparql_column_order:
        if col in all_keys and col not in excluded_keys:
            headers.append(col)
    
    # Add any other columns that might exist but weren't in the expected order
    # (shouldn't happen, but just in case)
    for key in sorted(all_keys):
        if key not in headers and key not in excluded_keys and key not in ['views_sv', 'views_fi', 'views_en', 'length_sv', 'length_fi', 'length_en']:
            headers.append(key)
    
    # Add pageview columns at the end
    headers.extend(['length_sv', 'length_fi', 'length_en', 'views_sv', 'views_fi', 'views_en'])
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # If this is the 'item' column, make it a hyperlink to Wikidata
            if header == 'item' and value:
                qcode = extract_qcode(value)
                if qcode:
                    wiki_url = f"https://www.wikidata.org/wiki/{qcode}"
                    cell.hyperlink = wiki_url
                    cell.font = Font(color="0000FF", underline="single")
                    # Store Q-code as display value
                    cell.value = qcode
            
            # If this is a Wikipedia title column, make it a hyperlink to Wikipedia
            elif header in ['wp_sv_title', 'wp_fi_title', 'wp_en_title'] and value:
                # Extract language code from column name (wp_sv_title -> sv, etc.)
                lang_code = header.replace('wp_', '').replace('_title', '')
                # Create Wikipedia URL - Wikipedia uses underscores for spaces in URLs
                # First replace spaces with underscores, then URL-encode any remaining special chars
                title_url = value.replace(' ', '_')
                # URL-encode the title (but keep underscores)
                title_url_encoded = quote(title_url, safe='_')
                wiki_url = f"https://{lang_code}.wikipedia.org/wiki/{title_url_encoded}"
                cell.hyperlink = wiki_url
                cell.font = Font(color="0000FF", underline="single")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row_idx in range(2, len(rows) + 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    wb.save(output_file)
    print(f"Excel file saved to {output_file}")


sparql_query = """
SELECT DISTINCT 
?item ?dob 
# ?pobLabel ?dod ?podLabel
?itemLabel_sv ?itemLabel_fi ?itemLabel_en
?itemDescription_sv ?itemDescription_fi ?itemDescription_en
?forfattare_sida ?forfattare_ref 
?wp_sv_title ?wp_fi_title ?wp_en_title
WHERE { 
  ?item p:P1343 ?s. 
  ?s ps:P1343 wd:Q136677319 . 
  ?s pq:P304 ?forfattare_sida.
  OPTIONAL{?s pq:P958 ?forfattare_ref.}
  OPTIONAL { ?item wdt:P31 ?inst }
  OPTIONAL { ?item wdt:P569 ?dob }
  OPTIONAL { ?item wdt:P570 ?dod }
  OPTIONAL { ?item wdt:P19 ?pob }
  OPTIONAL { ?item wdt:P20 ?pod }
  
  # Fetch item labels in specific languages
  OPTIONAL { ?item rdfs:label ?itemLabel_sv. FILTER(LANG(?itemLabel_sv) = "sv") }
  OPTIONAL { ?item rdfs:label ?itemLabel_fi. FILTER(LANG(?itemLabel_fi) = "fi") }
  OPTIONAL { ?item rdfs:label ?itemLabel_en. FILTER(LANG(?itemLabel_en) = "en") }
  
  # Fetch item descriptions in specific languages
  OPTIONAL { ?item schema:description ?itemDescription_sv. FILTER(LANG(?itemDescription_sv) = "sv") }
  OPTIONAL { ?item schema:description ?itemDescription_fi. FILTER(LANG(?itemDescription_fi) = "fi") }
  OPTIONAL { ?item schema:description ?itemDescription_en. FILTER(LANG(?itemDescription_en) = "en") }

  OPTIONAL { ?wp_sv_url schema:about ?item . 
             ?wp_sv_url schema:isPartOf <https://sv.wikipedia.org/>; 
             schema:name ?wp_sv_title. }
  OPTIONAL { ?wp_fi_url schema:about ?item . 
             ?wp_fi_url schema:isPartOf <https://fi.wikipedia.org/>; 
             schema:name ?wp_fi_title. }
  OPTIONAL { ?wp_en_url schema:about ?item . 
             ?wp_en_url schema:isPartOf <https://en.wikipedia.org/>; 
             schema:name ?wp_en_title. }
SERVICE wikibase:label { bd:serviceParam wikibase:language "sv,fi,en". }
} ORDER BY ?forfattare_ref
# LIMIT 10
"""


def main():
    print("=" * 60)
    print("Wikipedia Pageview Statistics Fetcher")
    print("=" * 60)
    print()
    
    # Get date range for the past year
    start_date, end_date = get_date_range_one_year_back()
    print(f"Fetching pageviews from {start_date} to {end_date}")
    print(f"Time period: Last 12 months")
    print()
    
    # Run SPARQL query
    bindings = run_sparql_query(sparql_query)
    
    if not bindings:
        print("No results found from SPARQL query")
        return
    
    # Process SPARQL results into rows
    all_rows = []
    for binding in bindings:
        # Convert binding to a more usable format
        row = {}
        for key, value_obj in binding.items():
            if isinstance(value_obj, dict) and 'value' in value_obj:
                row[key] = value_obj['value']
            else:
                row[key] = ''
        all_rows.append(row)
    
    # Remove duplicate rows based on forfattare_ref, keeping the first occurrence
    seen_refs = set()
    unique_rows = []
    duplicates_removed = 0
    
    for row in all_rows:
        forfattare_ref = row.get('forfattare_ref', '')
        # If forfattare_ref is empty or None, treat it as empty string and still allow it
        if forfattare_ref not in seen_refs:
            seen_refs.add(forfattare_ref)
            unique_rows.append(row)
        else:
            duplicates_removed += 1
    
    if duplicates_removed > 0:
        print(f"Removed {duplicates_removed} duplicate row(s) based on 'forfattare_ref'")
        print(f"Keeping {len(unique_rows)} unique row(s) for pageview fetching\n")
    
    # Now fetch pageviews only for unique rows
    total = len(unique_rows)
    rows = []
    
    for idx, row in enumerate(unique_rows, 1):
        # Get Wikipedia titles
        wp_sv_title = clean_title(row.get('wp_sv_title', ''))
        wp_fi_title = clean_title(row.get('wp_fi_title', ''))
        wp_en_title = clean_title(row.get('wp_en_title', ''))
        
        print(f"\n[{idx}/{total}] Processing: {row.get('itemLabel_sv', row.get('itemLabel_en', 'Unknown'))}")
        
        # Fetch pageviews for each language
        views_sv = 0
        views_fi = 0
        views_en = 0
        # Fetch article length for each language
        length_sv = 0
        length_fi = 0
        length_en = 0
        
        if wp_sv_title:
            print(f"  Fetching pageviews for sv.wikipedia.org/wiki/{wp_sv_title}")
            views_sv = get_wikipedia_pageviews(wp_sv_title, 'sv', start_date, end_date)
            print(f"    Total views: {views_sv:,}")
            length_sv = get_wikipedia_article_length(wp_sv_title, 'sv')
            print(f"    Length (chars): {length_sv:,}")
            time.sleep(0.2)  # Be nice to the API
        
        if wp_fi_title:
            print(f"  Fetching pageviews for fi.wikipedia.org/wiki/{wp_fi_title}")
            views_fi = get_wikipedia_pageviews(wp_fi_title, 'fi', start_date, end_date)
            print(f"    Total views: {views_fi:,}")
            length_fi = get_wikipedia_article_length(wp_fi_title, 'fi')
            print(f"    Length (chars): {length_fi:,}")
            time.sleep(0.2)
        
        if wp_en_title:
            print(f"  Fetching pageviews for en.wikipedia.org/wiki/{wp_en_title}")
            views_en = get_wikipedia_pageviews(wp_en_title, 'en', start_date, end_date)
            print(f"    Total views: {views_en:,}")
            length_en = get_wikipedia_article_length(wp_en_title, 'en')
            print(f"    Length (chars): {length_en:,}")
            time.sleep(0.2)
        
        # Add article lengths and pageview counts to row
        row['length_sv'] = length_sv
        row['length_fi'] = length_fi
        row['length_en'] = length_en
        row['views_sv'] = views_sv
        row['views_fi'] = views_fi
        row['views_en'] = views_en
        
        rows.append(row)
    
    # Create Excel output
    output_file = '05_output.xlsx'
    print(f"\n\nCreating Excel file with {len(rows)} rows...")
    create_excel(rows, output_file)
    
    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()